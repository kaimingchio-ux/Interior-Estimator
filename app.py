import streamlit as st
import google.generativeai as genai
from PIL import Image as PilImage
import json
import pandas as pd
import io
import plotly.express as px
from datetime import datetime
import os 

# Excel 排版專用套件
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage 

# ==========================================
# 1. 網頁基本設定
# ==========================================
st.set_page_config(page_title="夜間部設計-室內裝修估價系統", layout="wide")

LOGO_FILE = "資產 6.png"
if os.path.exists(LOGO_FILE):
    st.image(LOGO_FILE, width=180)

st.title("夜間部設計-室內裝修估價系統")
st.markdown("上傳各空間的『現況照』與『參考照』,讓我們系統幫你精準估價!")

# ==========================================
# 2. 側邊欄：設定 API Key
# ==========================================
selected_model_name = None
with st.sidebar:
    st.header("⚙️ 系統設定")
    api_key = st.text_input("請輸入你的 Gemini API Key", type="password")
    if api_key:
        try:
            genai.configure(api_key=api_key)
            available_models = [m.name.replace('models/', '') for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and ('vision' in m.name or '1.5' in m.name or 'pro' in m.name)]
            if available_models:
                st.success("✅ API 連線成功！")
                selected_model_name = st.selectbox("🤖 請選擇引擎", available_models)
        except Exception as e:
            st.error(f"❌ 錯誤：{e}")

# ==========================================
# 3. 系統記憶體初始化
# ==========================================
if "room_count" not in st.session_state: st.session_state.room_count = 1
if "quote_df" not in st.session_state: st.session_state.quote_df = pd.DataFrame()
if "db_df" not in st.session_state:
    st.session_state.db_df = pd.DataFrame([
        {"Category": "拆除工程", "Item": "隔間牆拆除", "Unit": "坪", "Price": 12000},
        {"Category": "木作工程", "Item": "平釘天花板", "Unit": "坪", "Price": 4500},
        {"Category": "油漆工程", "Item": "全室乳膠漆", "Unit": "坪", "Price": 1800}
    ])

# ==========================================
# 4. 頁籤切換
# ==========================================
tab_est, tab_db = st.tabs(["🏠 估價作業區", "📚 專屬價格資料庫"])

with tab_db:
    st.subheader("📝 價目表編輯")
    uploaded_db_file = st.file_uploader("📥 上傳 Excel 價目表", type=["xlsx", "xls", "csv"])
    if uploaded_db_file:
        st.session_state.db_df = pd.read_csv(uploaded_db_file) if uploaded_db_file.name.endswith('.csv') else pd.read_excel(uploaded_db_file)
    edited_db = st.data_editor(st.session_state.db_df, key="db_editor", num_rows="dynamic", use_container_width=True)

with tab_est:
    def add_room(): st.session_state.room_count += 1
    st.button("➕ 新增空間", on_click=add_room)
    
    project_data = []
    for i in range(st.session_state.room_count):
        with st.expander(f"📍 空間 {i+1}", expanded=True):
            c1, c2 = st.columns(2)
            with c1:
                r_name = st.text_input(f"名稱", key=f"n_{i}")
                r_req = st.text_area(f"需求", key=f"r_{i}")
            with c2:
                b_imgs = st.file_uploader(f"現況", type=['png','jpg','jpeg'], accept_multiple_files=True, key=f"b_{i}")
                a_imgs = st.file_uploader(f"參考", type=['png','jpg','jpeg'], accept_multiple_files=True, key=f"a_{i}")
            project_data.append({"name": r_name, "req": r_req, "before": [PilImage.open(img) for img in b_imgs] if b_imgs else [], "after": [PilImage.open(img) for img in a_imgs] if a_imgs else []})

    if st.button("🚀 開始 AI 估價", type="primary"):
        if not api_key: st.error("⚠️ 請輸入 API Key")
        else:
            with st.spinner("🧠 AI 分析中..."):
                try:
                    db_csv = edited_db.to_csv(index=False)
                    contents = ["設計師估價單。優先參考價格庫：\n"+db_csv, "格式：[{\"Category\": \"工種\", \"Item\": \"項目\", \"Qty\": 1, \"Unit\": \"單位\", \"Price\": 1000}]"]
                    for room in project_data:
                        contents.append(f"空間：{room['name']}\n需求：{room['req']}")
                        contents.extend(room['before']); contents.extend(room['after'])
                    model = genai.GenerativeModel(selected_model_name)
                    res = model.generate_content(contents)
                    
                    raw_text = res.text
                    start_idx = raw_text.find('[')
                    end_idx = raw_text.rfind(']') + 1
                    
                    if start_idx != -1 and end_idx != -1:
                        json_str = raw_text[start_idx:end_idx]
                        st.session_state.quote_df = pd.DataFrame(json.loads(json_str))
                        st.session_state.quote_df['Qty'] = pd.to_numeric(st.session_state.quote_df.get('Qty', 0), errors='coerce').fillna(0)
                        st.session_state.quote_df['Price'] = pd.to_numeric(st.session_state.quote_df.get('Price', 0), errors='coerce').fillna(0)
                        st.session_state.quote_df['Total'] = st.session_state.quote_df['Qty'] * st.session_state.quote_df['Price']
                    else:
                        st.error("⚠️ 資料解析失敗！")
                except Exception as e: st.error(f"❌ 錯誤：{e}")

    # ==========================================
    # 報價單顯示與下載按鈕區
    # ==========================================
    if not st.session_state.quote_df.empty:
        st.subheader("📋 估價明細表")
        edited_quote = st.data_editor(st.session_state.quote_df, key="q_editor", num_rows="dynamic", use_container_width=True)
        edited_quote['Total'] = edited_quote['Qty'].astype(float) * edited_quote['Price'].astype(float)
        
        st.markdown("---")
        total_val = edited_quote['Total'].sum()
        summary_df = edited_quote.groupby('Category')['Total'].sum().reset_index()
        summary_df = summary_df[summary_df['Total'] > 0]

        # 網頁端甜甜圈圖
        st.markdown("<h2 style='text-align: center;'>📊 裝潢預算構成比例圖</h2>", unsafe_allow_html=True)
        fig = px.pie(summary_df, values='Total', names='Category', hole=0.55, color_discrete_sequence=px.colors.qualitative.Pastel)
        fig.update_traces(textposition='outside', textinfo='percent+label', texttemplate='%{label}<br>%{percent}')
        fig.update_layout(
            showlegend=False, 
            annotations=[dict(text=f"總預算 NT$<br><span style='font-size:32px; color:#555;'><b>{total_val:,.0f}</b></span>", x=0.5, y=0.5, font_size=20, showarrow=False)], 
            height=650, margin=dict(t=40, b=100, l=40, r=40) 
        )
        st.plotly_chart(fig, use_container_width=True)

        col1, col_center, col3 = st.columns([1, 2, 1])
        with col_center:
            st.markdown(f"<div style='text-align: center; padding: 20px; border: 2px solid #f0f2f6; border-radius: 15px;'><h3>💰 預估總費用</h3><h1 style='color: #E63946;'>NT$ {total_val:,.0f}</h1></div>", unsafe_allow_html=True)
            st.write("") 

            def generate_styled_excel(df, s_data):
                wb = Workbook()
                ws1 = wb.active; ws1.title = "商務報價單"
                ws2 = wb.create_sheet("預算分析圖")
                
                f_title = Font(name='微軟正黑體', size=16, bold=True)
                f_head = Font(name='微軟正黑體', size=11, bold=True, color="FFFFFF")
                f_bold = Font(name='微軟正黑體', size=11, bold=True)
                f_norm = Font(name='微軟正黑體', size=11)
                
                fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                fill_grey = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                align_c = Alignment(horizontal='center', vertical='center')
                thin = Side(border_style="thin"); border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

                def set_b(ws, rng):
                    for row in ws[rng]:
                        for cell in row: cell.border = border_all

                # ----------------------------------------------------
                # 第一頁報價單
                # ----------------------------------------------------
                ws1.column_dimensions['A'].width = 15; ws1.column_dimensions['B'].width = 40
                ws1.column_dimensions['C'].width = 10; ws1.column_dimensions['D'].width = 15
                ws1.column_dimensions['E'].width = 15; ws1.column_dimensions['F'].width = 18

                ws1['A2'] = "夜間部設計 Nocturne Design Studio"; ws1['A2'].font = f_title; ws1['A2'].alignment = align_c
                ws1.merge_cells('A2:F3'); set_b(ws1, 'A2:F3')
                ws1['E4'] = "報價單編號：" + datetime.now().strftime("%Y%m%d%H%M"); ws1.merge_cells('E4:F4')
                ws1['E5'] = "報價日期：" + datetime.now().strftime("%Y-%m-%d"); ws1.merge_cells('E5:F5')
                ws1['A7'] = " 客戶名稱：                  聯絡人：                  聯絡電話："; ws1['A7'].font = f_bold
                ws1.merge_cells('A7:F7'); set_b(ws1, 'A7:F7')

                headers = ["工種 (Category)", "項目內容 (Item)", "數量", "單位", "單價", "費用 (Total)"]
                for i, h in enumerate(headers, 1):
                    c = ws1.cell(9, i, h); c.font = f_head; c.fill = fill_black; c.alignment = align_c; c.border = border_all

                curr_r = 10
                for _, r in df.iterrows():
                    ws1.cell(curr_r, 1, r['Category']).font = f_norm; ws1.cell(curr_r, 1).alignment = align_c
                    ws1.cell(curr_r, 2, r['Item']).font = f_norm
                    ws1.cell(curr_r, 3, r['Qty']).alignment = align_c; ws1.cell(curr_r, 4, r['Unit']).alignment = align_c
                    ws1.cell(curr_r, 5, r['Price']).number_format = '#,##0'
                    ws1.cell(curr_r, 6, f"=C{curr_r}*E{curr_r}").number_format = '#,##0'
                    set_b(ws1, f'A{curr_r}:F{curr_r}'); curr_r += 1
                
                # 合計區塊
                start_c = 10; end_c = curr_r - 1
                ws1.cell(curr_r, 1, "合計 Subtotal").alignment = align_c; ws1.cell(curr_r, 1).fill = fill_grey
                ws1.cell(curr_r, 6, f"=SUM(F{start_c}:F{end_c})").number_format = '#,##0'; ws1.cell(curr_r, 6).fill = fill_grey
                ws1.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=5); set_b(ws1, f'A{curr_r}:F{curr_r}'); curr_r += 1
                
                ws1.cell(curr_r, 1, "5%稅金 VAT").alignment = align_c; ws1.cell(curr_r, 1).fill = fill_grey
                ws1.cell(curr_r, 6, f"=F{curr_r-1}*0.05").number_format = '#,##0'; ws1.cell(curr_r, 6).fill = fill_grey
                ws1.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=5); set_b(ws1, f'A{curr_r}:F{curr_r}'); curr_r += 1
                
                ws1.cell(curr_r, 1, "總計 Total").alignment = align_c; ws1.cell(curr_r, 1).font = f_bold; ws1.cell(curr_r, 1).fill = fill_grey
                ws1.cell(curr_r, 6, f"=F{curr_r-2}+F{curr_r-1}").number_format = '#,##0'; ws1.cell(curr_r, 6).font = f_bold; ws1.cell(curr_r, 6).fill = fill_grey
                ws1.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=5); set_b(ws1, f'A{curr_r}:F{curr_r}')

                curr_r += 2
                ws1.cell(curr_r, 1, "合作備註：\n1. 若有任何約定條款，請於簽訂本估價單時一併提出。\n2. 付款方式分為：訂金30%、工程款40%、尾款30%。\n3. 本報價單有效期限為30天。").alignment = Alignment(vertical='top', wrap_text=True)
                ws1.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r+3, end_column=6); set_b(ws1, f'A{curr_r}:F{curr_r+3}')

                curr_r += 5
                ws1.cell(curr_r, 1, "專案負責人簽章：").font = f_bold; ws1.cell(curr_r, 4, "客戶簽章：").font = f_bold
                curr_r += 2
                ws1.cell(curr_r, 1, "_____________________"); ws1.cell(curr_r, 4, "_____________________")

                ws1.page_setup.paperSize = ws1.PAPERSIZE_A4; ws1.page_setup.orientation = ws1.ORIENTATION_PORTRAIT; ws1.page_setup.fitToPage = True; ws1.page_setup.fitToWidth = 1; ws1.page_setup.fitToHeight = 0; ws1.print_options.horizontalCentered = True; ws1.page_margins.left = 0.5; ws1.page_margins.right = 0.5; ws1.page_margins.top = 0.75; ws1.page_margins.bottom = 0.75

                # ----------------------------------------------------
                # 第二頁：完美雙拼排版
                # ----------------------------------------------------
                ws2.column_dimensions['A'].width = 5   
                ws2.column_dimensions['B'].width = 25  
                ws2.column_dimensions['C'].width = 18  
                ws2.column_dimensions['D'].width = 15  
                ws2.column_dimensions['E'].width = 5   
                ws2.column_dimensions['F'].width = 92  
                
                ws2['B4'] = "裝修預算分項金額表"; ws2['B4'].font = f_title; ws2['B4'].alignment = align_c; ws2.merge_cells('B4:D4')
                ws2['F4'] = "裝修預算構成比例圖"; ws2['F4'].font = f_title; ws2['F4'].alignment = align_c

                if not s_data.empty:
                    s_data['%'] = s_data['Total'] / total_val
                    s_data = s_data.sort_values('Total', ascending=False)

                    start_tbl = 6
                    for i, h in enumerate(["工項名稱 (Category)", "預估金額 (Amount)", "佔比 (Percent)"], 2): 
                        c = ws2.cell(start_tbl, i, h); c.font = f_head; c.fill = fill_black; c.alignment = align_c; c.border = border_all

                    tbl_row = start_tbl + 1
                    for _, r in s_data.iterrows():
                        ws2.cell(tbl_row, 2, r['Category']).alignment = align_c; ws2.cell(tbl_row, 2).border = border_all
                        ws2.cell(tbl_row, 3, r['Total']).number_format = '#,##0'; ws2.cell(tbl_row, 3).alignment = align_c; ws2.cell(tbl_row, 3).border = border_all
                        ws2.cell(tbl_row, 4, r['%']).number_format = '0.0%'; ws2.cell(tbl_row, 4).alignment = align_c; ws2.cell(tbl_row, 4).border = border_all
                        tbl_row += 1

                    ws2.cell(tbl_row, 2, "總計 Total").font = f_bold; ws2.cell(tbl_row, 2).fill = fill_grey; ws2.cell(tbl_row, 2).alignment = align_c; ws2.cell(tbl_row, 2).border = border_all
                    ws2.cell(tbl_row, 3, total_val).font = f_bold; ws2.cell(tbl_row, 3).number_format = '#,##0'; ws2.cell(tbl_row, 3).fill = fill_grey; ws2.cell(tbl_row, 3).alignment = align_c; ws2.cell(tbl_row, 3).border = border_all
                    ws2.cell(tbl_row, 4, 1.0).font = f_bold; ws2.cell(tbl_row, 4).number_format = '0.0%'; ws2.cell(tbl_row, 4).fill = fill_grey; ws2.cell(tbl_row, 4).alignment = align_c; ws2.cell(tbl_row, 4).border = border_all
                    
                    # 🔥 直接產生圖片，不再需要 Checkbox 判斷
                    excel_fig = px.pie(s_data, values='Total', names='Category', hole=0.55, color_discrete_sequence=px.colors.qualitative.Pastel)
                    excel_fig.update_traces(textposition='outside', textinfo='percent+label', texttemplate='%{label}<br>%{percent}', marker=dict(line=dict(color='#FFFFFF', width=2)))
                    excel_fig.update_layout(
                        paper_bgcolor='white', plot_bgcolor='white', showlegend=False, 
                        annotations=[dict(text=f"總預算<br><span style='font-size:32px; color:#555;'><b>{total_val:,.0f}</b></span>", x=0.5, y=0.5, font_size=20, showarrow=False)],
                        width=800, height=600, margin=dict(t=40, b=80, l=150, r=150), font=dict(family="Microsoft JhengHei")
                    )
                    
                    try:
                        img_bytes = excel_fig.to_image(format="png", scale=2)
                        chart_img = OpenpyxlImage(io.BytesIO(img_bytes)); chart_img.width = 650; chart_img.height = 480
                        ws2.add_image(chart_img, 'F6') 
                    except Exception as e: pass
                    
                ws2.page_setup.paperSize = ws2.PAPERSIZE_A4; ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE; ws2.page_setup.fitToPage = True; ws2.page_setup.fitToWidth = 1; ws2.page_setup.fitToHeight = 1; ws2.print_options.horizontalCentered = True; ws2.page_margins.left = 0.5; ws2.page_margins.right = 0.5; ws2.page_margins.top = 0.5; ws2.page_margins.bottom = 0.5
                
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()

            try:
                # 🌟 直接執行，無須傳入 Checkbox 變數
                excel_bin = generate_styled_excel(edited_quote, summary_df)
                st.download_button("📥 下載品牌究極版報價單 (A4 列印版)", excel_bin, f"夜間部設計裝修工程報價單_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
            except Exception as e: st.error(f"❌ Excel 產生失敗：{e}")